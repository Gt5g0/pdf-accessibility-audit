"""Build the color-coded Excel workbook for audit results.

The workbook has a Summary sheet plus one detail sheet per PDF. This module
contains all openpyxl-specific formatting so the rest of the pipeline can remain
focused on discovery, validation, and parsing.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List, Set

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from validation_report import ValidationReport

_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER_FONT = Font(bold=True)


class ReportBuilder:
    """Construct the Summary sheet and one detail sheet per PDF."""

    _SHEET_INVALID = re.compile(r'[/\\*?\[\]:]')

    @classmethod
    def _unique_sheet_name(cls, student: str, file_name: str, used: Set[str]) -> str:
        """Return an Excel-safe, unique sheet title and record it in ``used``.

        Excel sheet titles have character and length constraints; this helper
        sanitizes and deduplicates names to avoid export failures.
        """
        s = (student or "root").strip() or "root"
        f = (file_name or "file").strip() or "file"
        s = cls._SHEET_INVALID.sub("_", s)
        f = cls._SHEET_INVALID.sub("_", f)
        sep = " || "
        max_len = 31
        if len(s) + len(sep) + len(f) <= max_len:
            base = s + sep + f
        else:
            budget = max_len - len(sep)
            left = max(1, budget // 2)
            s_part = s[:left]
            f_part = f[: max(1, budget - len(s_part))]
            base = (s_part + sep + f_part)[:max_len]

        name = base[:max_len]
        n = 2
        while name in used:
            suffix = "_%d" % n
            name = (base[: max_len - len(suffix)] + suffix)[:max_len]
            n += 1
        used.add(name)
        return name

    def build_and_save(self, reports: List[ValidationReport], output_path: Path) -> Path:
        """Write the audit workbook to ``output_path`` and return the saved path."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        headers = [
            "Student Name",
            "File Name",
            "Checkpoints Passed",
            "Checkpoints Failed",
            "Overall",
        ]
        for col, h in enumerate(headers, start=1):
            cell = summary.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 60
        summary.column_dimensions["C"].width = 17
        summary.column_dimensions["D"].width = 17

        used_names: Set[str] = {"Summary"}

        for row_idx, rep in enumerate(reports, start=2):
            passed = sum(1 for c in rep.checkpoints if c.passed)
            failed = sum(1 for c in rep.checkpoints if not c.passed)
            overall_txt = "PASS" if rep.overall_pass else "FAIL"

            summary.cell(row=row_idx, column=1, value=rep.pdf_record.student or "")
            summary.cell(row=row_idx, column=2, value=rep.pdf_record.name)
            summary.cell(row=row_idx, column=3, value=passed)
            summary.cell(row=row_idx, column=4, value=failed)
            oc = summary.cell(row=row_idx, column=5, value=overall_txt)
            oc.fill = _PASS_FILL if rep.overall_pass else _FAIL_FILL

        for rep in reports:
            sheet_name = self._unique_sheet_name(rep.pdf_record.student, rep.pdf_record.name, used_names)
            ws = wb.create_sheet(title=sheet_name)

            ws.cell(row=1, column=1, value="Student Name").font = _HEADER_FONT
            ws.cell(row=1, column=2, value=rep.pdf_record.student or "")
            ws.cell(row=2, column=1, value="File Name").font = _HEADER_FONT
            ws.cell(row=2, column=2, value=rep.pdf_record.name)
            ws.cell(row=3, column=1, value="Folder Path").font = _HEADER_FONT
            ws.cell(row=3, column=2, value=rep.pdf_record.path)
            ws.cell(row=4, column=1, value="Overall").font = _HEADER_FONT
            ov = ws.cell(row=4, column=2, value="PASS" if rep.overall_pass else "FAIL")
            ov.fill = _PASS_FILL if rep.overall_pass else _FAIL_FILL

            hdr_row = 6
            detail_headers = [
                "Checkpoint ID",
                "Description",
                "Status",
                "Failures",
                "Pages",
            ]
            for col, h in enumerate(detail_headers, start=1):
                c = ws.cell(row=hdr_row, column=col, value=h)
                c.font = _HEADER_FONT
            ws.column_dimensions["A"].width = 13
            ws.column_dimensions["B"].width = 130

            for i, cp in enumerate(rep.checkpoints, start=1):
                r = hdr_row + i
                ws.cell(row=r, column=1, value=cp.id)
                ws.cell(row=r, column=2, value=cp.description)
                st = ws.cell(row=r, column=3, value="PASS" if cp.passed else "FAIL")
                ws.cell(row=r, column=4, value=cp.failed_occurrences)
                ws.cell(row=r, column=5, value=cp.failed_occurrence_pages or "")
                fill = _PASS_FILL if cp.passed else _FAIL_FILL
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = fill

        wb.save(output_path)
        return output_path
